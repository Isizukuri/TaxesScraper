import requests
from BeautifulSoup import BeautifulSoup as BS
from openpyxl import load_workbook
import csv

url = 'http://apps.irs.gov/app/stdc/stdc.html'
years = [2005, 2006, 2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014]
incomeRanges = [x for x in xrange(1, 20)]
exemptions = [x for x in xrange(1, 7)]

wb = load_workbook(filename='zip_code_database.xlsx', read_only=True)
ws = wb['zip_code_database']
rows = ws['A2':'A42522']
zips = []
for row in rows:
    buf = str(row[0].value)
    if len(buf) < 5:
        buf = ((5-len(buf)))*'0'+buf
    zips.append(buf)

csvFile = open('results.csv', 'a')
writer = csv.writer(csvFile)
writer.writerow(('Year','Income Range', 'exemptions', 'Move', 'ZIP Code', 'City or County ID', 'State Tax',
                    'Local Tax', 'Percents', 'State Tax Amount', 'Local Tax Amount', 'Total Tax'))
csvFile.close()

for Year in years:
    for incomeRange in incomeRanges:
        for exemption in exemptions:
            for zipInfo in zips:

                s = requests.Session()
                get = s.get(url)

                page1_parameters = {'_page': 0, '_target1': 'Continue'}



                page2_parameters = {'_page': 1, '_target2': 'Continue',
                                    'selectedYear': Year}


                page3_parameters = {'_page': 2, '_target3': 'Continue',
                                    'incomeRange': incomeRange, 'exemptions': exemption}

                
                    
                page4_parameters = {'_page': 3, '_target4': 'Continue',
                                    'initialZipInfo.zip': zipInfo}

                post1 = s.post(url, data=page1_parameters)
                post2= s.post(url, data=page2_parameters)
                post3 = s.post(url, data=page3_parameters)
                post4 = s.post(url, data=page4_parameters)

                post4Soup = BS(post4.text)

                if post4Soup.findAll(color='red'):
                    print 'Invalid ZIP: '+str(zipInfo)
                    pass

                else:
                
                    rows = post4Soup.findAll(type='radio')
                    idList = [str(row['value']) for row in rows]
                    idList = idList[:-1]

                    countyId = idList[0]

                    for countyId in idList:
                        s = requests.Session()
                        get = s.get(url)

                        page1_parameters = {'_page': 0, '_target1': 'Continue'}

                        
                        page2_parameters = {'_page': 1, '_target2': 'Continue',
                                            'selectedYear': Year}

                        
                        page3_parameters = {'_page': 2, '_target3': 'Continue',
                                            'incomeRange': incomeRange, 'exemptions': exemption}

                        
                        page4_parameters = {'_page': 3, '_target4': 'Continue',
                                            'initialZipInfo.zip': zipInfo}

                        post1 = s.post(url, data=page1_parameters)
                        post2= s.post(url, data=page2_parameters)
                        post3 = s.post(url, data=page3_parameters)
                        post4 = s.post(url, data=page4_parameters)
                        page5_parameters = {'_page': 4, '_target5': 'Continue',
                                            'initialZipInfo.cityCountyId': countyId}

                        post5 = s.post(url, data=page5_parameters)

                        page6_parameters = {'_page': 5, '_target6': 'Continue',
                                            'didYouMove': 'false'}

                        post6 = s.post(url, data=page6_parameters)

                        page7_parameters = {'_page': 10, '_target11': 'Continue'}

                        post7 = s.post(url, data=page7_parameters)

                        resultPage = BS(post7.text)


                        results = {}
                        results['incomeRange'] = resultPage.table.find('table').findAll('tr')[3].find(colspan='7').string.replace('&#036;','$ ')
                        results['exemptions'] = resultPage.table.find('table').findAll('tr')[4].find(colspan='7').string
                        results['moveData'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[0].string
                        results['zipCode'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[1].string
                        results['cityCountyState'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[2].string
                        results['stateTax'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[3].string
                        results['localTax'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[4].string
                        results['percentage'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[5].string
                        results['stateTaxAmount'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[6].string.replace('&#036;','$ ')
                        results['localTaxAmount'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[7].string.replace('&#036;','$ ')
                        results['totalTax'] = resultPage.table.find('table').findAll('tr')[7].findAll('td')[8].string.replace('&#036;','$ ')

                        csvFile = open('results.csv', 'a')
                        writer = csv.writer(csvFile)
                        writer.writerow((str(Year),results['incomeRange'],  results['exemptions'], results['moveData'], results['zipCode'],
                                         results['cityCountyState'], results['stateTax'], results['localTax'], results['percentage'], results['stateTaxAmount'],
                                         results['localTaxAmount'], results['totalTax']))
                        
                        csvFile.close()
                        print str(Year)+' '+zipInfo+' '+str(countyId)+' '+'Row Succes'
                        

